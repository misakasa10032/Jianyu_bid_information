import os
import sys
from selenium import webdriver
import time
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
import string
import zipfile

def create_proxyauth_extension(proxy_host, proxy_port,
                               proxy_username, proxy_password,
                               scheme='http', plugin_path=None):
    """代理认证插件
 
    args:
        proxy_host (str): 你的代理地址或者域名（str类型）
        proxy_port (int): 代理端口号（int类型）
        proxy_username (str):用户名（字符串）
        proxy_password (str): 密码 （字符串）
    kwargs:
        scheme (str): 代理方式 默认http
        plugin_path (str): 扩展的绝对路径
 
    return str -> plugin_path
    """
    
 
    if plugin_path is None:
        plugin_path = 'vimm_chrome_proxyauth_plugin.zip'
 
    manifest_json = """
    {
        "version": "1.0.0",
        "manifest_version": 2,
        "name": "Chrome Proxy",
        "permissions": [
            "proxy",
            "tabs",
            "unlimitedStorage",
            "storage",
            "<all_urls>",
            "webRequest",
            "webRequestBlocking"
        ],
        "background": {
            "scripts": ["background.js"]
        },
        "minimum_chrome_version":"22.0.0"
    }
    """
 
    background_js = string.Template(
    """
    var config = {
            mode: "fixed_servers",
            rules: {
              singleProxy: {
                scheme: "${scheme}",
                host: "${host}",
                port: parseInt(${port})
              },
              bypassList: ["foobar.com"]
            }
          };
 
    chrome.proxy.settings.set({value: config, scope: "regular"}, function() {});
 
    function callbackFn(details) {
        return {
            authCredentials: {
                username: "${username}",
                password: "${password}"
            }
        };
    }
 
    chrome.webRequest.onAuthRequired.addListener(
                callbackFn,
                {urls: ["<all_urls>"]},
                ['blocking']
    );
    """
    ).substitute(
        host=proxy_host,
        port=proxy_port,
        username=proxy_username,
        password=proxy_password,
        scheme=scheme,
    )
    with zipfile.ZipFile(plugin_path, 'w') as zp:
        zp.writestr("manifest.json", manifest_json)
        zp.writestr("background.js", background_js)
 
    return plugin_path

proxyauth_plugin_path_1 = create_proxyauth_extension(
    proxy_host = "47.92.113.149",
    proxy_port = 16819,
    proxy_username = "m18362928852",
    proxy_password = "fyrxlezu"
)

arg1 = '--proxy-server=http://218.60.8.99:3129'
arg2 = '--proxy-server=http://218.60.8.98:3129'
arg3 = '--proxy-server=http://140.143.96.216:80'
arg4 = '--proxy-server=http://39.135.35.17:80'
arg5 = '--proxy-server=http://39.135.35.18:80'
arg6 = '--proxy-server=http://39.135.35.16:80'
arg7 = '--proxy-server=http://39.135.35.19:80'
arg8 = '--proxy-server=http://182.92.68.60:3128'


url = 'https://www.jianyu360.com/jylab/supsearch/index.html'
xlsx_set = ['C:/Users/3.5/Desktop/远程医疗.xlsx', 'C:/Users/3.5/Desktop/健康信息.xlsx']
driver = webdriver.Chrome()
chrome_options = webdriver.ChromeOptions()
for address in xlsx_set:
	try:
		wb = load_workbook(address)
		wb_sheet = wb.get_sheet_names()
		wb2 = Workbook()
		new_name = (address.split('/')[-1]).replace('.xlsx', '_amended.xlsx')
		for sn in wb_sheet:
			ws = wb.get_sheet_by_name(sn)
			wb2.create_sheet(sn)
			ws2 = wb2.get_sheet_by_name(sn)
			ws2.cell(row = 1, column = 10).value = '地区'
			start_row = 3
			for i in range(start_row, ws.max_row + 1):
				residue = (i - start_row + 1) % 64
				if	(residue == 1 or residue == 0 or residue == 3):
					driver.close()
					chrome_options.add_argument(arg1)
					driver = webdriver.Chrome(chrome_options = chrome_options)
				if residue == 9:
					driver.close()
					chrome_options.add_argument(arg2)
					driver = webdriver.Chrome(chrome_options = chrome_options)
					driver.get(url)
				if residue == 17:
					driver.close()
					chrome_options.add_argument(arg3)
					driver = webdriver.Chrome(chrome_options = chrome_options)
					driver.get(url)
				if residue == 25:
					driver.close()
					chrome_options.add_argument(arg4)
					driver = webdriver.Chrome(chrome_options = chrome_options)
					driver.get(url)
				if residue == 33:
					driver.close()
					chrome_options.add_argument(arg5)
					driver = webdriver.Chrome(chrome_options = chrome_options)
					driver.get(url)
				if residue == 41:
					driver.close()
					chrome_options.add_argument(arg6)
					driver = webdriver.Chrome(chrome_options = chrome_options)
					driver.get(url)
				if residue == 49:
					driver.close()
					chrome_options.add_argument(arg7)
					driver = webdriver.Chrome(chrome_options = chrome_options)
					driver.get(url)
				if residue == 57:
					driver.close()
					chrome_options.add_argument(arg8)
					driver = webdriver.Chrome(chrome_options = chrome_options)
					driver.get(url)
				keyword = ws.cell(row = i, column = 2).value
				try:
					driver.find_element_by_id('searchinput').send_keys(keyword)
				except:
					driver.close()
					chrome_options.add_argument(arg8)
					driver = webdriver.Chrome(chrome_options = chrome_options)
					driver.get(url)
					driver.find_element_by_id('searchinput').send_keys(keyword)
				try:
					driver.find_element_by_xpath('//*[@id="zbSeatchT"]/input[2]').send_keys(Keys.ENTER)
				except:
					driver.close()
					continue
				try:
					driver.find_element_by_xpath('//*[@id="searchInner"]/div[2]/div[3]/div[3]').get_attribute('style')
				except:
					driver.close()
					continue
				if driver.find_element_by_xpath('//*[@id="searchInner"]/div[2]/div[3]/div[3]').get_attribute('style') != 'display: none;' :
					areas = driver.find_element_by_xpath('//*[@id="searchInner"]/div[2]/div[3]/div[3]/div[1]/ul/li[1]/div/div[2]/a[1]').get_attribute('textContent')
				else:
					areas = ' '
				ws2.cell(row = i, column = 10).value = areas
			for i in range(1, ws.max_row + 1):
				for j in range(1, ws.max_column + 1):
					ws2.cell(row = i, column = j).value = ws.cell(row = i, column = j).value
		wb2.save(new_name)
		print(new_name + ' has been finished')
	except:
		wb2.save(new_name)
		print(new_name + ' has been failed')